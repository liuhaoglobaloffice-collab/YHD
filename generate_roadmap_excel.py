#!/usr/bin/env python3
"""
鎏灏 AI-OS 规划路线图 Excel 生成器
生成完整的项目规划和时间线表格
"""

from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def create_roadmap_excel():
    """创建鎏灏规划路线图 Excel 表格"""
    
    wb = Workbook()
    
    # 删除默认的 Sheet
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # 1. 总览表
    create_overview_sheet(wb)
    
    # 2. 16周详细计划表
    create_16week_plan_sheet(wb)
    
    # 3. Phase 1-3 详细任务表
    create_phase1_sheet(wb)
    create_phase2_sheet(wb)
    create_phase3_sheet(wb)
    
    # 4. 功能清单表
    create_features_sheet(wb)
    
    # 5. 里程碑表
    create_milestones_sheet(wb)
    
    # 6. 资源需求表
    create_resources_sheet(wb)
    
    # 7. 技术栈表
    create_tech_stack_sheet(wb)
    
    # 保存文件
    filename = f"鎏灏AI-OS_完整规划路线图_{datetime.now().strftime('%Y%m%d')}.xlsx"
    wb.save(filename)
    print(f"✅ Excel 表格已生成: {filename}")
    return filename


def create_overview_sheet(wb):
    """创建总览表"""
    ws = wb.create_sheet("📋 总览")
    
    # 设置列宽
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 15
    
    # 标题
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=14)
    
    ws['A1'] = "项目信息"
    ws['A1'].fill = header_fill
    ws['A1'].font = header_font
    ws.merge_cells('A1:C1')
    
    # 项目基本信息
    data = [
        ["项目名称", "鎏灏 AI-OS (LiuHao AI Operating System)", ""],
        ["项目愿景", "企业级 AI 操作系统，AI 原生企业转型平台", ""],
        ["当前版本", "Y1.0 (v1.0.0)", ""],
        ["开发阶段", "Week 3 Day 3", "60%"],
        ["启动日期", "2026-08-22", ""],
        ["预计完成", "2026-12-06 (16周)", ""],
        ["", "", ""],
        ["规划版本", "修正说明", "状态"],
        ["12周 FINAL版", "删除桌面+移动+粤语（已废弃）", "❌ 错误"],
        ["16周修正版", "保留桌面+移动+粤语（当前执行）", "✅ 执行中"],
    ]
    
    for row_idx, row_data in enumerate(data, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            
            # 表头样式
            if row_idx == 9:
                cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
                cell.font = Font(bold=True)
    
    # 核心目标
    ws['A13'] = "核心目标"
    ws['A13'].fill = header_fill
    ws['A13'].font = header_font
    ws.merge_cells('A13:C13')
    
    goals = [
        ["🎯 统一 AI 接入", "整合 6 大 AI 智能体 (GPT-4o, Grok, Claude, DeepSeek, Gemini, Kimi)", "✅"],
        ["🔐 企业级安全", "RBAC 权限、JWT 认证、审计日志、数据加密", "✅"],
        ["🧠 智能决策", "AI 驱动的业务分析、自动化工作流、CEO 决策仪表板", "🔄"],
        ["🌐 多平台支持", "Web + 桌面 (Electron) + 移动 (React Native)", "📅"],
        ["🗣️ 粤语全栈", "粤语 TTS、ASR、NLP、AI 对话", "📅"],
        ["📈 业务赋能", "供应商智能、CRM、销售漏斗、运营报表", "🔄"],
    ]
    
    for row_idx, goal in enumerate(goals, start=14):
        for col_idx, value in enumerate(goal, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)


def create_16week_plan_sheet(wb):
    """创建16周详细计划表"""
    ws = wb.create_sheet("📅 16周计划")
    
    # 设置列宽
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 50
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 12
    
    # 表头
    headers = ["周次", "阶段/模块", "主要任务", "交付成果", "状态"]
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # 计划数据
    plans = [
        # Phase 1
        ["Phase 1", "核心价值验证 (Week 2-8)", "", "", ""],
        ["Week 2", "供应商智能数据层", "供应商数据模型、CRUD、AI采集、风险评估、Dashboard API", "供应商智能系统", "80%"],
        ["Week 3", "API完善与测试加固", "统一错误处理、API文档、集成测试、性能优化", "测试覆盖率≥85%", "60%"],
        ["Week 4", "前端项目搭建", "React+TS+Vite、TailwindCSS、赛博朋克主题、组件库、路由系统", "前端脚手架+组件库", "0%"],
        ["Week 5", "CEO Dashboard", "实时仪表板、KPI卡片、趋势图表、AI员工管理、任务中心", "CEO Dashboard", "0%"],
        ["Week 6", "供应商管理前端", "供应商列表/详情、筛选排序、表单、风险评估可视化", "供应商管理界面", "0%"],
        ["Week 7", "前端完善与优化", "响应式适配、E2E测试、性能优化、用户体验优化", "前端优化完成", "0%"],
        ["Week 8", "Phase 1 集成测试", "前后端联调、E2E测试、性能压测、Demo准备", "Phase 1 可演示", "0%"],
        
        # Phase 2
        ["Phase 2", "系统能力增强 (Week 9-13)", "", "", ""],
        ["Week 9", "10个AI员工(上)", "销售部门3个、供应链部门3个", "6个AI员工", "0%"],
        ["Week 10", "10个AI员工(下)", "运营部门4个、协作系统、工作流增强", "10个AI员工+协作", "0%"],
        ["Week 11", "销售漏斗自动化", "销售漏斗模型、数据可视化、AI辅助销售", "销售漏斗系统", "0%"],
        ["Week 12", "简化CRM系统", "客户管理、商机管理、CRM前端集成", "CRM系统", "0%"],
        ["Week 13", "运营报表系统", "核心报表、报表自动化、数据大屏", "运营报表", "0%"],
        
        # Phase 3
        ["Phase 3", "多平台扩展 (Week 14-16)", "", "", ""],
        ["Week 14", "桌面应用 (Electron)", "Electron搭建、原生功能、桌面特性、打包测试", "Win/macOS 应用", "0%"],
        ["Week 15", "移动应用 (React Native)", "RN搭建、核心页面适配、原生功能、打包测试", "Android/iOS 应用", "0%"],
        ["Week 16", "粤语全栈支持", "粤语TTS、粤语ASR、粤语NLP、语言检测", "粤语全栈", "0%"],
        
        # Phase 4
        ["Week 17", "最终发布（可选）", "全系统集成测试、Bug修复、生产部署、UAT、发布", "Y1.0 发布", "0%"],
    ]
    
    # 填充数据
    for row_idx, plan in enumerate(plans, start=2):
        for col_idx, value in enumerate(plan, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            
            # Phase 标题样式
            if plan[0].startswith("Phase"):
                cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
                cell.font = Font(bold=True, size=11)
            
            # 状态颜色
            if col_idx == 5:  # 状态列
                if value == "80%" or value == "60%":
                    cell.fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
                elif value == "0%":
                    cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")


def create_phase1_sheet(wb):
    """创建 Phase 1 详细任务表"""
    ws = wb.create_sheet("Phase 1 - 核心验证")
    
    # 设置列宽
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 60
    ws.column_dimensions['D'].width = 12
    
    # 表头
    headers = ["周次 Day", "模块", "详细任务", "状态"]
    header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Week 2 任务
    tasks = [
        ["Week 2", "", "", ""],
        ["Day 1-3", "供应商数据模型", "- 4张表设计 (Supplier, SupplierContact, SupplierCertificate, SupplierRisk)\n- CRUD 服务开发 (20+方法)\n- 单元测试 (25个)", "✅"],
        ["Day 1-3", "AI 数据采集", "- SupplierDataCollector Agent 开发\n- 数据抓取与清洗\n- 85%+ 准确率验证", "✅"],
        ["Day 4", "风险评估", "- 6维度评估引擎\n- 风险等级分类 (LOW/MEDIUM/HIGH/CRITICAL)\n- 评估算法开发", "⏳"],
        ["Day 4", "Dashboard API", "- 6个统计端点\n- 实时数据聚合\n- 响应优化", "⏳"],
        ["Day 5", "演示数据", "- 生成50+供应商\n- Week 2 总结报告\n- Week 3 准备", "📅"],
        
        ["Week 3", "", "", ""],
        ["Day 1-2", "API 完善", "- 统一错误处理机制\n- API 文档完善 (100%)\n- 响应格式标准化", "✅"],
        ["Day 3", "集成测试", "- Business API 测试覆盖\n- 整体测试通过率 92.3%\n- 测试报告生成", "✅"],
        ["Day 4-5", "测试加固", "- 异常场景测试\n- E2E 测试修复\n- 覆盖率提升至 85%+", "📅"],
        ["Day 6-7", "性能优化", "- API 响应时间优化 (<200ms)\n- 数据库查询优化\n- Week 3 总结", "📅"],
        
        ["Week 4", "", "", ""],
        ["Day 1-2", "前端初始化", "- React 18 + TypeScript + Vite\n- TailwindCSS 配置\n- 赛博朋克主题设计", "📅"],
        ["Day 3-5", "组件库", "- 20+ 基础组件\n- 赛博朋克风格\n- 组件文档", "📅"],
        ["Day 6-7", "路由系统", "- 4级菜单导航\n- DashboardLayout 布局\n- 权限路由守卫", "📅"],
        
        ["Week 5", "", "", ""],
        ["Day 1-3", "实时仪表板", "- 6个 KPI 卡片\n- 4个趋势图表 (ECharts)\n- 最新动态列表", "📅"],
        ["Day 4-5", "AI员工管理", "- AI员工列表页\n- 员工详情页\n- 绩效展示", "📅"],
        ["Day 6-7", "任务中心", "- 任务列表\n- 任务详情\n- 任务创建/编辑", "📅"],
        
        ["Week 6", "", "", ""],
        ["Day 1-3", "供应商列表", "- 供应商列表页\n- 供应商详情页\n- 筛选/排序/搜索", "📅"],
        ["Day 4-5", "供应商表单", "- 创建/编辑表单\n- 联系人/证书管理\n- 表单验证", "📅"],
        ["Day 6-7", "风险评估UI", "- 风险评估卡片\n- 风险雷达图\n- 风险趋势图", "📅"],
        
        ["Week 7", "", "", ""],
        ["Day 1-3", "UX 优化", "- 响应式适配\n- 加载动画\n- 错误提示优化", "📅"],
        ["Day 4-5", "E2E 测试", "- Playwright 测试\n- 组件测试\n- 页面测试", "📅"],
        ["Day 6-7", "性能优化", "- 代码分割\n- 懒加载\n- 打包优化", "📅"],
        
        ["Week 8", "", "", ""],
        ["Day 1-2", "前后端联调", "- API 集成调试\n- 错误处理\n- 权限控制测试", "📅"],
        ["Day 3-4", "E2E 测试", "- 完整流程测试\n- 跨模块测试\n- 回归测试", "📅"],
        ["Day 5", "性能压测", "- 100并发测试\n- 性能基准\n- 优化建议", "📅"],
        ["Day 6-7", "Demo 准备", "- 演示环境搭建\n- 演示数据准备\n- Phase 1 总结报告", "📅"],
    ]
    
    for row_idx, task in enumerate(tasks, start=2):
        for col_idx, value in enumerate(task, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            
            # Week 标题样式
            if value.startswith("Week"):
                cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
                cell.font = Font(bold=True)
            
            # 设置行高
            ws.row_dimensions[row_idx].height = 45 if "\n" in str(value) else 20


def create_phase2_sheet(wb):
    """创建 Phase 2 详细任务表"""
    ws = wb.create_sheet("Phase 2 - 能力增强")
    
    # 设置列宽
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 60
    ws.column_dimensions['D'].width = 12
    
    # 表头
    headers = ["周次 Day", "模块", "详细任务", "状态"]
    header_fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    tasks = [
        ["Week 9", "", "", ""],
        ["Day 1-2", "销售部门AI", "- 销售经理 AI (SalesManagerAgent)\n- 客户开发 AI (CustomerDevelopmentAgent)\n- 商机分析 AI (OpportunityAnalysisAgent)", "📅"],
        ["Day 3-4", "供应链AI", "- 供应商分析 AI (SupplierAnalysisAgent)\n- 采购建议 AI (ProcurementAdvisorAgent)\n- 物流优化 AI (LogisticsOptimizerAgent)", "📅"],
        ["Day 5-7", "测试集成", "- 单元测试 (5+/Agent)\n- AI 调用测试\n- 前端集成", "📅"],
        
        ["Week 10", "", "", ""],
        ["Day 1-3", "运营部门AI", "- 数据分析 AI (DataAnalystAgent)\n- 报表生成 AI (ReportGeneratorAgent)\n- 风险监控 AI (RiskMonitorAgent)\n- 决策支持 AI (DecisionSupportAgent)", "📅"],
        ["Day 4-5", "协作系统", "- AI 员工路由算法\n- 任务自动分配\n- 结果传递机制", "📅"],
        ["Day 6-7", "前端集成", "- AI 员工详情页更新\n- 协作流程可视化\n- 任务执行监控", "📅"],
        
        ["Week 11", "", "", ""],
        ["Day 1-3", "销售漏斗", "- Lead → Opportunity → Quote → Deal\n- 阶段转化规则\n- 自动推进逻辑", "📅"],
        ["Day 4-5", "数据可视化", "- 漏斗图 (ECharts)\n- 转化率分析\n- 销售预测", "📅"],
        ["Day 6-7", "AI 辅助", "- 线索评分\n- 跟进建议\n- 话术推荐", "📅"],
        
        ["Week 12", "", "", ""],
        ["Day 1-3", "客户管理", "- 客户列表 (CRUD)\n- 客户详情页\n- 客户分类与标签\n- 沟通记录", "📅"],
        ["Day 4-5", "商机管理", "- 商机创建/编辑\n- 商机阶段管理\n- 客户关联", "📅"],
        ["Day 6-7", "CRM 前端", "- 客户管理页面\n- 商机管理页面\n- 数据展示", "📅"],
        
        ["Week 13", "", "", ""],
        ["Day 1-3", "核心报表", "- 销售报表 (日/周/月)\n- 供应商报表\n- AI 员工效能\n- 财务概览", "📅"],
        ["Day 4-5", "报表自动化", "- 定时生成\n- 邮件推送\n- PDF 导出", "📅"],
        ["Day 6-7", "数据大屏", "- 实时数据大屏\n- 全屏展示\n- TV 模式", "📅"],
    ]
    
    for row_idx, task in enumerate(tasks, start=2):
        for col_idx, value in enumerate(task, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            
            if value.startswith("Week"):
                cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
                cell.font = Font(bold=True)
            
            ws.row_dimensions[row_idx].height = 45 if "\n" in str(value) else 20


def create_phase3_sheet(wb):
    """创建 Phase 3 详细任务表"""
    ws = wb.create_sheet("Phase 3 - 多平台扩展")
    
    # 设置列宽
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 60
    ws.column_dimensions['D'].width = 12
    
    # 表头
    headers = ["周次 Day", "模块", "详细任务", "状态"]
    header_fill = PatternFill(start_color="C65911", end_color="C65911", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    tasks = [
        ["Week 14", "桌面应用 (Electron)", "", ""],
        ["Day 1-2", "项目搭建", "- Electron + React 集成\n- 主进程/渲染进程架构\n- IPC 通信机制\n- 开发环境配置", "📅"],
        ["Day 3-4", "原生功能", "- 系统托盘 (Tray)\n- 原生菜单 (Menu)\n- 原生通知 (Notification)\n- 全局快捷键\n- 窗口管理", "📅"],
        ["Day 5-6", "桌面特性", "- 开机自启动\n- 最小化到托盘\n- 窗口置顶模式\n- 截图功能\n- 本地存储 (SQLite)", "📅"],
        ["Day 7", "打包测试", "- Windows 打包 (electron-builder)\n- macOS 打包\n- 自动更新配置\n- 桌面端测试", "📅"],
        
        ["Week 15", "移动应用 (React Native)", "", ""],
        ["Day 1-2", "RN 搭建", "- React Native 项目初始化\n- React Navigation 配置\n- 移动端主题适配\n- 状态管理 (Zustand)", "📅"],
        ["Day 3-4", "核心页面", "- 登录页\n- CEO Dashboard (移动版)\n- 供应商列表 (移动版)\n- 任务列表 (移动版)", "📅"],
        ["Day 5-6", "原生功能", "- 推送通知 (Firebase/极光)\n- 相机拍照\n- 图片选择\n- 文件上传\n- 语音录制", "📅"],
        ["Day 7", "打包测试", "- Android APK 打包\n- iOS IPA 打包 (可选)\n- 移动端测试\n- 应用商店准备", "📅"],
        
        ["Week 16", "粤语全栈支持", "", ""],
        ["Day 1-2", "粤语 TTS", "- VITS 粤语模型集成\n- 支持广州话/香港话\n- 男女声可选\n- 语速/音调调节\n- TTS API 封装", "📅"],
        ["Day 3-4", "粤语 ASR", "- Whisper Large-v3 基础\n- 粤语专有词库 (1000+词)\n- 俚语/歇后语处理\n- 后处理矫正\n- ASR API 封装", "📅"],
        ["Day 5-6", "粤语 NLP", "- 粤语优化 Prompt\n- 500+ 核心词汇词典\n- 100+ 俚语/歇后语库\n- 中英混杂支持 (港式粤语)\n- 粤语对话测试", "📅"],
        ["Day 7", "语言检测", "- 粤语特征词检测\n- 自动切换粤语模式\n- 混合对话支持\n- UI 繁简切换\n- Week 16 总结", "📅"],
    ]
    
    for row_idx, task in enumerate(tasks, start=2):
        for col_idx, value in enumerate(task, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            
            if value.startswith("Week"):
                cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
                cell.font = Font(bold=True)
            
            ws.row_dimensions[row_idx].height = 45 if "\n" in str(value) else 20


def create_features_sheet(wb):
    """创建功能清单表"""
    ws = wb.create_sheet("🎯 功能清单")
    
    # 设置列宽
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 50
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 15
    
    # 表头
    headers = ["#", "功能模块", "详细说明", "优先级", "状态"]
    header_fill = PatternFill(start_color="7030A0", end_color="7030A0", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    features = [
        [1, "统一 AI 接入", "6 大 AI 智能体：GPT-4o, Grok, Claude, DeepSeek, Gemini, Kimi", "P0", "✅ 完成"],
        [2, "企业级安全", "RBAC 权限、JWT 认证、审计日志、数据加密", "P0", "✅ 完成"],
        [3, "用户管理", "用户注册、登录、权限分配、多角色支持", "P0", "✅ 完成"],
        [4, "AI 智能路由", "智能选择最合适的 AI 模型、负载均衡", "P0", "✅ 完成"],
        [5, "记忆系统", "短期记忆、长期记忆、上下文管理", "P0", "✅ 完成"],
        [6, "公司大脑", "实体管理、事实管理、关系管理、知识沉淀", "P0", "✅ 完成"],
        [7, "任务管理", "任务创建、分配、跟踪、完成", "P0", "✅ 完成"],
        [8, "工作流引擎", "线性工作流、任务编排、自动化执行", "P0", "✅ 完成"],
        [9, "供应商智能", "数据采集、风险评估、智能推荐", "P0", "⏳ 80%"],
        [10, "CEO Dashboard", "实时仪表板、KPI 展示、数据可视化", "P0", "📅 Week 5"],
        [11, "10个 AI 员工", "销售3个、供应链3个、运营4个", "P1", "📅 Week 9-10"],
        [12, "销售漏斗", "Lead→Opportunity→Quote→Deal 自动化", "P1", "📅 Week 11"],
        [13, "简化 CRM", "客户管理、商机管理、沟通记录", "P1", "📅 Week 12"],
        [14, "运营报表", "销售报表、供应商报表、效能报表、数据大屏", "P1", "📅 Week 13"],
        [15, "桌面应用", "Electron、系统托盘、原生通知、自动更新", "P1", "📅 Week 14"],
        [16, "移动应用", "React Native、Android APK、iOS IPA、推送通知", "P1", "📅 Week 15"],
        [17, "粤语全栈", "粤语 TTS、粤语 ASR、粤语 NLP、语言检测", "P1", "📅 Week 16"],
        ["", "", "", "", ""],
        ["已删除功能（延后到 Y1.1）", "", "", "", ""],
        ["-", "多租户系统", "租户隔离、Token 池管理", "P2", "❌ 延后"],
        ["-", "Token 隐秘调度", "隐秘 Token 调度策略", "P2", "❌ 延后"],
        ["-", "营销模块", "营销自动化、活动管理", "P2", "❌ 延后"],
        ["-", "研发模块", "研发管理、版本控制", "P2", "❌ 延后"],
        ["-", "本地 LLM", "Ollama、Qwen2.5、pgvector、RAG", "P2", "❌ 延后"],
        ["-", "元认知层", "自我反思、自我进化、无限进化", "P2", "❌ 延后"],
        ["-", "32 专家", "32个专业 AI 员工（简化为10个）", "P2", "❌ 简化"],
    ]
    
    for row_idx, feature in enumerate(features, start=2):
        for col_idx, value in enumerate(feature, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            
            # 标题行样式
            if isinstance(feature[0], str) and "已删除" in str(feature[0]):
                cell.fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
                cell.font = Font(bold=True, color="FFFFFF")


def create_milestones_sheet(wb):
    """创建里程碑表"""
    ws = wb.create_sheet("🏁 里程碑")
    
    # 设置列宽
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 50
    ws.column_dimensions['D'].width = 15
    
    # 表头
    headers = ["时间", "里程碑", "交付成果", "状态"]
    header_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    header_font = Font(bold=True, size=11)
    
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # 计算日期
    start_date = datetime(2026, 8, 22)
    milestones = [
        [start_date + timedelta(weeks=0), "🚀 项目启动", "项目初始化、架构搭建、基础设施", "✅ 完成"],
        [start_date + timedelta(weeks=2), "M1: 供应商智能", "供应商智能系统、风险评估、Dashboard API", "⏳ 80%"],
        [start_date + timedelta(weeks=3), "M2: API 完善", "测试覆盖率 ≥85%、API 文档完整、性能优化", "⏳ 60%"],
        [start_date + timedelta(weeks=8), "M3: Phase 1 完成", "Web 端可演示、前后端集成、系统可用", "📅 2026-10-17"],
        [start_date + timedelta(weeks=13), "M4: Phase 2 完成", "10个 AI 员工、销售漏斗、CRM、运营报表", "📅 2026-11-21"],
        [start_date + timedelta(weeks=14), "M5: 桌面应用", "Windows/macOS 桌面应用、原生功能集成", "📅 2026-11-28"],
        [start_date + timedelta(weeks=15), "M6: 移动应用", "Android/iOS 移动应用、推送通知", "📅 2026-12-05"],
        [start_date + timedelta(weeks=16), "M7: 粤语全栈", "粤语 TTS/ASR/NLP、语言检测、繁简切换", "📅 2026-12-06"],
        [start_date + timedelta(weeks=17), "🎉 Y1.0 发布", "全系统集成、生产部署、正式发布", "📅 2026-12-13"],
    ]
    
    for row_idx, milestone in enumerate(milestones, start=2):
        for col_idx, value in enumerate(milestone, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            
            # 格式化日期
            if col_idx == 1 and isinstance(value, datetime):
                cell.value = value.strftime("%Y-%m-%d")
            else:
                cell.value = value
            
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            
            # 状态颜色
            if col_idx == 4:
                if "完成" in str(value):
                    cell.fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
                elif "%" in str(value):
                    cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
                elif "📅" in str(value):
                    cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")


def create_resources_sheet(wb):
    """创建资源需求表"""
    ws = wb.create_sheet("💰 资源需求")
    
    # 设置列宽
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 50
    ws.column_dimensions['D'].width = 15
    
    # 人力资源
    ws['A1'] = "人力资源"
    ws['A1'].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws['A1'].font = Font(bold=True, color="FFFFFF", size=12)
    ws.merge_cells('A1:D1')
    
    human_resources = [
        ["角色", "人数", "职责", "时间"],
        ["技术负责人", "1", "架构设计、技术决策、代码审查", "全程 (16周)"],
        ["后端开发", "2", "API 开发、数据库设计、业务逻辑", "16周"],
        ["前端开发", "2", "UI/UX、组件开发、前端集成", "Week 4-17"],
        ["AI 工程师", "1", "AI 模型集成、Prompt 优化、粤语 NLP", "Week 9-16"],
        ["测试工程师", "1", "测试用例、自动化测试、性能测试", "Week 3-17"],
        ["UI/UX 设计师", "1", "界面设计、交互设计、视觉设计", "Week 4-15"],
        ["产品经理", "1", "需求管理、项目协调、用户验收", "全程 (16周)"],
    ]
    
    for row_idx, resource in enumerate(human_resources, start=2):
        for col_idx, value in enumerate(resource, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            
            if row_idx == 2:
                cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
                cell.font = Font(bold=True)
    
    # 技术资源
    ws['A10'] = "技术资源"
    ws['A10'].fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    ws['A10'].font = Font(bold=True, color="FFFFFF", size=12)
    ws.merge_cells('A10:D10')
    
    tech_resources = [
        ["类型", "配置", "说明", "预估成本"],
        ["开发机", "8台", "16GB+ RAM, SSD, Windows/macOS", "-"],
        ["测试服务器", "2台", "32GB RAM, PostgreSQL + Redis", "$200/月"],
        ["生产服务器", "4台", "64GB RAM, 负载均衡", "$500/月"],
        ["数据库服务器", "2台", "128GB RAM, PostgreSQL 主从", "$300/月"],
        ["Redis 集群", "3台", "32GB RAM", "$150/月"],
        ["AI API", "-", "GPT-4o, Grok, Claude, DeepSeek, Gemini, Kimi", "$3000-5000/月"],
        ["云服务", "-", "CDN, 存储, 域名, SSL", "$500/月"],
    ]
    
    for row_idx, resource in enumerate(tech_resources, start=11):
        for col_idx, value in enumerate(resource, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            
            if row_idx == 11:
                cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
                cell.font = Font(bold=True)
    
    # 预算汇总
    ws['A20'] = "预算汇总"
    ws['A20'].fill = PatternFill(start_color="C65911", end_color="C65911", fill_type="solid")
    ws['A20'].font = Font(bold=True, color="FFFFFF", size=12)
    ws.merge_cells('A20:D20')
    
    budget = [
        ["项目", "金额", "说明", "备注"],
        ["技术资源", "$24,000", "服务器+AI API (4个月)", ""],
        ["软件工具", "$2,000", "开发工具、设计软件、项目管理", ""],
        ["人力成本", "待定", "9人 × 4个月", "根据团队配置"],
        ["总预算", "待定", "人力成本 + $26,000", ""],
    ]
    
    for row_idx, item in enumerate(budget, start=21):
        for col_idx, value in enumerate(item, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            
            if row_idx == 21:
                cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
                cell.font = Font(bold=True)


def create_tech_stack_sheet(wb):
    """创建技术栈表"""
    ws = wb.create_sheet("🔧 技术栈")
    
    # 设置列宽
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 50
    
    # 表头
    headers = ["分类", "技术/工具", "说明"]
    header_fill = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    tech_stack = [
        ["后端框架", "Python 3.13", "主编程语言"],
        ["后端框架", "FastAPI", "高性能 Web 框架、异步 API"],
        ["后端框架", "SQLAlchemy", "ORM 数据库操作"],
        ["后端框架", "Pydantic", "数据验证、设置管理"],
        ["", "", ""],
        ["数据库", "PostgreSQL 15+", "生产环境主数据库"],
        ["数据库", "SQLite", "测试环境数据库"],
        ["数据库", "Redis 7+", "缓存、会话存储"],
        ["", "", ""],
        ["AI 服务", "OpenAI GPT-4o", "通用智能、代码生成"],
        ["AI 服务", "xAI Grok", "实时数据、推理能力"],
        ["AI 服务", "Anthropic Claude", "长文本理解、推理"],
        ["AI 服务", "DeepSeek", "中文优化、代码能力"],
        ["AI 服务", "Google Gemini", "多模态理解"],
        ["AI 服务", "Moonshot Kimi", "超长上下文 (128K tokens)"],
        ["", "", ""],
        ["安全", "JWT (HS256)", "Token 认证"],
        ["安全", "bcrypt", "密码哈希"],
        ["安全", "Fernet", "数据加密"],
        ["安全", "slowapi", "API 限流"],
        ["", "", ""],
        ["前端框架", "React 18", "UI 框架"],
        ["前端框架", "TypeScript", "类型安全"],
        ["前端框架", "Vite", "构建工具"],
        ["前端框架", "TailwindCSS", "样式框架"],
        ["前端框架", "ECharts", "数据可视化"],
        ["前端框架", "React Router", "路由管理"],
        ["", "", ""],
        ["桌面应用", "Electron 28+", "跨平台桌面应用"],
        ["桌面应用", "electron-builder", "打包工具"],
        ["桌面应用", "electron-updater", "自动更新"],
        ["", "", ""],
        ["移动应用", "React Native 0.73+", "跨平台移动应用"],
        ["移动应用", "React Navigation 6+", "移动端路由"],
        ["移动应用", "Firebase", "推送通知 (Android)"],
        ["移动应用", "APNs", "推送通知 (iOS)"],
        ["", "", ""],
        ["粤语支持", "VITS", "粤语 TTS (语音合成)"],
        ["粤语支持", "Whisper Large-v3", "粤语 ASR (语音识别)"],
        ["粤语支持", "自定义词库", "1000+ 粤语词汇、100+ 俚语"],
        ["", "", ""],
        ["测试", "pytest", "单元测试、集成测试"],
        ["测试", "pytest-asyncio", "异步测试"],
        ["测试", "Playwright", "E2E 测试"],
        ["测试", "pytest-cov", "测试覆盖率"],
        ["", "", ""],
        ["部署", "Docker", "容器化"],
        ["部署", "docker-compose", "服务编排"],
        ["部署", "Nginx", "反向代理、负载均衡"],
        ["部署", "GitHub Actions", "CI/CD"],
    ]
    
    for row_idx, tech in enumerate(tech_stack, start=2):
        for col_idx, value in enumerate(tech, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)


if __name__ == "__main__":
    print("🚀 开始生成鎏灏 AI-OS 规划路线图 Excel...")
    filename = create_roadmap_excel()
    print(f"✅ 完成！文件: {filename}")
    print(f"📊 包含以下工作表：")
    print("   1. 📋 总览")
    print("   2. 📅 16周计划")
    print("   3. Phase 1 - 核心验证")
    print("   4. Phase 2 - 能力增强")
    print("   5. Phase 3 - 多平台扩展")
    print("   6. 🎯 功能清单")
    print("   7. 🏁 里程碑")
    print("   8. 💰 资源需求")
    print("   9. 🔧 技术栈")
